# doplaty.py
# -*- coding: utf-8 -*-
"""
AIR/IRZ++ • Moduł dopłat (pełna wersja – część 1/4)
Autor: ChatGPT dla Feniksa
"""

import os
import sys
import sqlite3
import json
import datetime
import tkinter as tk
from tkinter import messagebox, filedialog
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from PIL import Image, ImageTk

# =============================
#   ŚCIEŻKI
# =============================
APPDIR = os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "KozyManager")
os.makedirs(APPDIR, exist_ok=True)

SESSION_FILE = os.path.join(APPDIR, "session_token.json")
DOPLATY_DB_FILE = os.path.join(APPDIR, "doplaty.db")
UPRAWY_LOG = os.path.join(APPDIR, "logs", "uprawy.log")
os.makedirs(os.path.dirname(UPRAWY_LOG), exist_ok=True)

# =============================
#   SESJA
# =============================
def load_session_username():
    if not os.path.exists(SESSION_FILE):
        return None
    try:
        with open(SESSION_FILE, "r", encoding="utf-8") as f:
            token = json.load(f)
        return token.get("username")
    except Exception:
        return None

# =============================
#   DB INIT
# =============================
def init_doplaty_db():
    conn = sqlite3.connect(DOPLATY_DB_FILE)
    cur = conn.cursor()

    cur.execute("""CREATE TABLE IF NOT EXISTS doplaty (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nazwa TEXT,
        opis TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS stawki (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        doplata TEXT,
        stawka REAL
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS terminy (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        doplata TEXT,
        od DATE,
        do DATE
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS checklist (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        doplata TEXT,
        dokument TEXT
    )""")

    cur.execute("""CREATE TABLE IF NOT EXISTS doplaty_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        rok INTEGER,
        doplata TEXT,
        kwota REAL
    )""")

    conn.commit()
    conn.close()

init_doplaty_db()

# =============================
#   GŁÓWNE OKNO
# =============================
class DoplatyApp(tb.Window):
    def __init__(self, username: str):
        super().__init__(themename="darkly")
        self.title(f"AIR/IRZ++ – Dopłaty ({username})")
        self.geometry("1200x750")
        self.resizable(True, True)
        self.username = username

        # tło
        try:
            bg = Image.open("OK.jpg")
            self.bg_tk = ImageTk.PhotoImage(bg.resize((1200, 750)))
            bg_label = tk.Label(self, image=self.bg_tk)
            bg_label.place(x=0, y=0, relwidth=1, relheight=1)
        except Exception:
            self.configure(bg="#1f1f1f")

        # zakładki
        notebook = tb.Notebook(self, bootstyle="primary")
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        # główne sekcje
        self.tab_lista = tb.Frame(notebook)      # dopłaty
        self.tab_stawki = tb.Frame(notebook)     # stawki
        self.tab_terminy = tb.Frame(notebook)    # terminy
        self.tab_checklist = tb.Frame(notebook)  # dokumenty
        self.tab_calc = tb.Frame(notebook)       # kalkulator
        self.tab_historia = tb.Frame(notebook)   # historia
        self.tab_alerty = tb.Frame(notebook)     # alerty
        self.tab_raport = tb.Frame(notebook)     # raporty
        self.tab_excel = tb.Frame(notebook)      # excel
        self.tab_uprawy = tb.Frame(notebook)     # uprawy
        self.tab_analizy = tb.Frame(notebook)    # analizy
        self.tab_backup = tb.Frame(notebook)     # backup/ustawienia

        notebook.add(self.tab_lista, text="Dopłaty")
        notebook.add(self.tab_stawki, text="Stawki")
        notebook.add(self.tab_terminy, text="Terminy")
        notebook.add(self.tab_checklist, text="Dokumenty")
        notebook.add(self.tab_calc, text="Kalkulator")
        notebook.add(self.tab_historia, text="Historia")
        notebook.add(self.tab_alerty, text="Alerty")
        notebook.add(self.tab_raport, text="Raporty")
        notebook.add(self.tab_excel, text="Excel")
        notebook.add(self.tab_uprawy, text="Uprawy")
        notebook.add(self.tab_analizy, text="Analizy")
        notebook.add(self.tab_backup, text="Backup/Ustawienia")

        # kolejne build_* będą w częściach 2–4
    # ------------------------------
    # 1. Dopłaty
    def build_tab_lista(self):
        tb.Label(self.tab_lista, text="Lista dopłat (edycja):", font="-size 12 -weight bold").pack(pady=10)
        self.lista_frame = tb.Frame(self.tab_lista)
        self.lista_frame.pack(fill="both", expand=True)
        self.refresh_lista_doplat()

        form = tb.Frame(self.tab_lista, padding=10)
        form.pack(fill="x")
        tb.Label(form, text="Nowa dopłata:").grid(row=0, column=0, padx=5)
        self.new_doplata = tb.Entry(form, width=30)
        self.new_doplata.grid(row=0, column=1, padx=5)
        tb.Button(form, text="Dodaj", bootstyle="success", command=self.add_doplata).grid(row=0, column=2, padx=5)

    def refresh_lista_doplat(self):
        for widget in self.lista_frame.winfo_children():
            widget.destroy()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT id,nazwa FROM doplaty")
        rows = cur.fetchall()
        conn.close()
        for fid, nazwa in rows:
            row_frame = tb.Frame(self.lista_frame)
            row_frame.pack(fill="x", pady=2)
            tb.Label(row_frame, text=nazwa).pack(side="left", padx=10)
            tb.Button(row_frame, text="Usuń", bootstyle="danger", command=lambda i=fid: self.delete_doplata(i)).pack(side="right")

    def add_doplata(self):
        nazwa = self.new_doplata.get().strip()
        if not nazwa: return
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("INSERT INTO doplaty(nazwa,opis) VALUES (?,?)", (nazwa,""))
        conn.commit()
        conn.close()
        self.new_doplata.delete(0,"end")
        self.refresh_lista_doplat()

    def delete_doplata(self, fid):
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("DELETE FROM doplaty WHERE id=?", (fid,))
        conn.commit()
        conn.close()
        self.refresh_lista_doplat()

    # ------------------------------
    # 2. Stawki
    def build_tab_stawki(self):
        tb.Label(self.tab_stawki, text="Stawki dopłat:", font="-size 12 -weight bold").pack(pady=10)
        self.stawki_frame = tb.Frame(self.tab_stawki)
        self.stawki_frame.pack(fill="both", expand=True)
        self.refresh_stawki()

        form = tb.Frame(self.tab_stawki, padding=10)
        form.pack(fill="x")
        tb.Label(form, text="Dopłata:").grid(row=0, column=0)
        self.stawka_doplata = tb.Entry(form, width=20)
        self.stawka_doplata.grid(row=0, column=1, padx=5)
        tb.Label(form, text="Stawka [zł/ha]:").grid(row=0, column=2)
        self.stawka_value = tb.Entry(form, width=10)
        self.stawka_value.grid(row=0, column=3, padx=5)
        tb.Button(form, text="Dodaj", bootstyle="success", command=self.add_stawka).grid(row=0, column=4, padx=5)

    def refresh_stawki(self):
        for widget in self.stawki_frame.winfo_children():
            widget.destroy()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT id,doplata,stawka FROM stawki")
        rows = cur.fetchall()
        conn.close()
        for fid, dop, stawka in rows:
            row_frame = tb.Frame(self.stawki_frame)
            row_frame.pack(fill="x", pady=2)
            tb.Label(row_frame, text=f"{dop}: {stawka} zł/ha").pack(side="left", padx=10)
            tb.Button(row_frame, text="Usuń", bootstyle="danger", command=lambda i=fid: self.delete_stawka(i)).pack(side="right")

    def add_stawka(self):
        dop = self.stawka_doplata.get().strip()
        try:
            val = float(self.stawka_value.get())
        except:
            return
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("INSERT INTO stawki(doplata,stawka) VALUES (?,?)", (dop,val))
        conn.commit()
        conn.close()
        self.stawka_doplata.delete(0,"end")
        self.stawka_value.delete(0,"end")
        self.refresh_stawki()

    def delete_stawka(self, fid):
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("DELETE FROM stawki WHERE id=?", (fid,))
        conn.commit()
        conn.close()
        self.refresh_stawki()

    # ------------------------------
    # 3. Terminy
    def build_tab_terminy(self):
        tb.Label(self.tab_terminy, text="Terminy naborów:", font="-size 12 -weight bold").pack(pady=10)
        self.term_frame = tb.Frame(self.tab_terminy)
        self.term_frame.pack(fill="both", expand=True)
        self.refresh_terminy()

        form = tb.Frame(self.tab_terminy, padding=10)
        form.pack(fill="x")
        tb.Label(form, text="Dopłata:").grid(row=0, column=0)
        self.term_doplata = tb.Entry(form, width=20)
        self.term_doplata.grid(row=0, column=1, padx=5)
        tb.Label(form, text="Od (YYYY-MM-DD):").grid(row=0, column=2)
        self.term_od = tb.Entry(form, width=12)
        self.term_od.grid(row=0, column=3, padx=5)
        tb.Label(form, text="Do (YYYY-MM-DD):").grid(row=0, column=4)
        self.term_do = tb.Entry(form, width=12)
        self.term_do.grid(row=0, column=5, padx=5)
        tb.Button(form, text="Dodaj", bootstyle="success", command=self.add_termin).grid(row=0, column=6, padx=5)

    def refresh_terminy(self):
        for widget in self.term_frame.winfo_children():
            widget.destroy()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT id,doplata,od,do FROM terminy")
        rows = cur.fetchall()
        conn.close()
        for fid, dop, od, do in rows:
            row_frame = tb.Frame(self.term_frame)
            row_frame.pack(fill="x", pady=2)
            tb.Label(row_frame, text=f"{dop}: {od} – {do}").pack(side="left", padx=10)
            tb.Button(row_frame, text="Usuń", bootstyle="danger", command=lambda i=fid: self.delete_termin(i)).pack(side="right")

    def add_termin(self):
        dop = self.term_doplata.get().strip()
        od = self.term_od.get().strip()
        do = self.term_do.get().strip()
        if not dop or not od or not do: return
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("INSERT INTO terminy(doplata,od,do) VALUES (?,?,?)", (dop,od,do))
        conn.commit()
        conn.close()
        self.term_doplata.delete(0,"end")
        self.term_od.delete(0,"end")
        self.term_do.delete(0,"end")
        self.refresh_terminy()

    def delete_termin(self, fid):
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("DELETE FROM terminy WHERE id=?", (fid,))
        conn.commit()
        conn.close()
        self.refresh_terminy()

    # ------------------------------
    # 4. Dokumenty
    def build_tab_checklist(self):
        tb.Label(self.tab_checklist, text="Dokumenty wymagane:", font="-size 12 -weight bold").pack(pady=10)
        self.doc_frame = tb.Frame(self.tab_checklist)
        self.doc_frame.pack(fill="both", expand=True)
        self.refresh_docs()

        form = tb.Frame(self.tab_checklist, padding=10)
        form.pack(fill="x")
        tb.Label(form, text="Dopłata:").grid(row=0, column=0)
        self.doc_doplata = tb.Entry(form, width=20)
        self.doc_doplata.grid(row=0, column=1, padx=5)
        tb.Label(form, text="Dokument:").grid(row=0, column=2)
        self.doc_name = tb.Entry(form, width=30)
        self.doc_name.grid(row=0, column=3, padx=5)
        tb.Button(form, text="Dodaj", bootstyle="success", command=self.add_doc).grid(row=0, column=4, padx=5)

    def refresh_docs(self):
        for widget in self.doc_frame.winfo_children():
            widget.destroy()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT id,doplata,dokument FROM checklist")
        rows = cur.fetchall()
        conn.close()
        for fid, dop, doc in rows:
            row_frame = tb.Frame(self.doc_frame)
            row_frame.pack(fill="x", pady=2)
            tb.Label(row_frame, text=f"{dop}: {doc}").pack(side="left", padx=10)
            tb.Button(row_frame, text="Usuń", bootstyle="danger", command=lambda i=fid: self.delete_doc(i)).pack(side="right")

    def add_doc(self):
        dop = self.doc_doplata.get().strip()
        doc = self.doc_name.get().strip()
        if not dop or not doc: return
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("INSERT INTO checklist(doplata,dokument) VALUES (?,?)", (dop,doc))
        conn.commit()
        conn.close()
        self.doc_doplata.delete(0,"end")
        self.doc_name.delete(0,"end")
        self.refresh_docs()

    def delete_doc(self, fid):
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("DELETE FROM checklist WHERE id=?", (fid,))
        conn.commit()
        conn.close()
        self.refresh_docs()
    # ------------------------------
    # 5. Kalkulator dopłat
    def build_tab_calc(self):
        tb.Label(self.tab_calc, text="Kalkulator dopłat:", font="-size 12 -weight bold").pack(pady=10)

        frame = tb.Frame(self.tab_calc, padding=20)
        frame.pack(fill="x")

        tb.Label(frame, text="Powierzchnia [ha]:").grid(row=0, column=0, sticky="e")
        self.kalk_area = tb.Entry(frame, width=10)
        self.kalk_area.grid(row=0, column=1, padx=5)

        tb.Label(frame, text="Rodzaj dopłaty:").grid(row=1, column=0, sticky="e")
        self.kalk_type = tb.Combobox(frame, width=25)
        self.kalk_type.grid(row=1, column=1, padx=5)
        self.refresh_calc_options()

        tb.Button(frame, text="Oblicz", bootstyle="success", command=self.calc_doplata).grid(row=2, column=0, columnspan=2, pady=10)

        self.kalk_result = tb.Label(frame, text="", font="-size 12 -weight bold")
        self.kalk_result.grid(row=3, column=0, columnspan=2, pady=10)

    def refresh_calc_options(self):
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT doplata FROM stawki")
        rows = [r[0] for r in cur.fetchall()]
        conn.close()
        self.kalk_type["values"] = rows

    def calc_doplata(self):
        try:
            area = float(self.kalk_area.get())
        except:
            messagebox.showerror("Błąd", "Podaj poprawną powierzchnię [ha].")
            return
        dtype = self.kalk_type.get()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT stawka FROM stawki WHERE doplata=?", (dtype,))
        row = cur.fetchone()
        conn.close()
        if not row:
            messagebox.showerror("Błąd", "Nie znaleziono stawki dla tej dopłaty.")
            return
        kwota = area * row[0]
        self.kalk_result.config(text=f"Kwota: {kwota:.2f} zł")

    # ------------------------------
    # 6. Historia wniosków
    def build_tab_historia(self):
        tb.Label(self.tab_historia, text="Historia wniosków:", font="-size 12 -weight bold").pack(pady=10)

        self.hist_frame = tb.Frame(self.tab_historia)
        self.hist_frame.pack(fill="both", expand=True)
        self.refresh_historia()

    def refresh_historia(self):
        for widget in self.hist_frame.winfo_children():
            widget.destroy()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT rok,doplata,kwota FROM doplaty_history WHERE username=?", (self.username,))
        rows = cur.fetchall()
        conn.close()
        if not rows:
            tb.Label(self.hist_frame, text="Brak danych.").pack()
        else:
            for r in rows:
                tb.Label(self.hist_frame, text=f"{r[0]} – {r[1]}: {r[2]} zł").pack(anchor="w", padx=20)

    # ------------------------------
    # 7. Alerty
    def build_tab_alerty(self):
        tb.Label(self.tab_alerty, text="Alerty (terminy naborów):", font="-size 12 -weight bold").pack(pady=10)

        self.alert_frame = tb.Frame(self.tab_alerty)
        self.alert_frame.pack(fill="both", expand=True)
        self.refresh_alerty()

    def refresh_alerty(self):
        for widget in self.alert_frame.winfo_children():
            widget.destroy()
        dzis = datetime.date.today()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT doplata, od, do FROM terminy")
        rows = cur.fetchall()
        conn.close()
        for dop, od, do in rows:
            try:
                end_date = datetime.datetime.strptime(do, "%Y-%m-%d").date()
                dni = (end_date - dzis).days
                if dni >= 0:
                    tb.Label(self.alert_frame, text=f"⏰ {dop}: do {do} (za {dni} dni)").pack(anchor="w", padx=20)
                else:
                    tb.Label(self.alert_frame, text=f"❌ {dop}: termin minął {do}").pack(anchor="w", padx=20)
            except:
                pass

    # ------------------------------
    # 8. Raporty
    def build_tab_raport(self):
        tb.Label(self.tab_raport, text="Raporty:", font="-size 12 -weight bold").pack(pady=10)
        tb.Button(self.tab_raport, text="Eksport CSV", bootstyle="success", command=self.export_csv).pack(pady=5)
        tb.Button(self.tab_raport, text="Eksport TXT", bootstyle="info", command=self.export_txt).pack(pady=5)

    def export_csv(self):
        fname = filedialog.asksaveasfilename(defaultextension=".csv")
        if not fname: return
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT rok,doplata,kwota FROM doplaty_history WHERE username=?", (self.username,))
        rows = cur.fetchall()
        conn.close()
        with open(fname, "w", encoding="utf-8") as f:
            f.write("Rok,Dopłata,Kwota\n")
            for r in rows:
                f.write(f"{r[0]},{r[1]},{r[2]}\n")
        messagebox.showinfo("Raport", "Zapisano plik CSV.")

    def export_txt(self):
        fname = filedialog.asksaveasfilename(defaultextension=".txt")
        if not fname: return
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT rok,doplata,kwota FROM doplaty_history WHERE username=?", (self.username,))
        rows = cur.fetchall()
        conn.close()
        with open(fname, "w", encoding="utf-8") as f:
            for r in rows:
                f.write(f"{r[0]} – {r[1]}: {r[2]} zł\n")
        messagebox.showinfo("Raport", "Zapisano plik TXT.")

    # ------------------------------
    # 9. Excel
    def build_tab_excel(self):
        tb.Label(self.tab_excel, text="Import/Eksport Excel:", font="-size 12 -weight bold").pack(pady=10)
        tb.Button(self.tab_excel, text="Eksport XLSX", bootstyle="success", command=self.export_xlsx).pack(pady=5)
        tb.Button(self.tab_excel, text="Import XLSX", bootstyle="info", command=self.import_xlsx).pack(pady=5)

    def export_xlsx(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Błąd", "Brak biblioteki openpyxl.")
            return
        fname = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if not fname: return
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT rok,doplata,kwota FROM doplaty_history WHERE username=?", (self.username,))
        rows = cur.fetchall()
        conn.close()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Rok", "Dopłata", "Kwota"])
        for r in rows:
            ws.append(r)
        wb.save(fname)
        messagebox.showinfo("Raport", "Zapisano plik XLSX.")

    def import_xlsx(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Błąd", "Brak biblioteki openpyxl.")
            return
        fname = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not fname: return
        wb = openpyxl.load_workbook(fname)
        ws = wb.active
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        for row in ws.iter_rows(min_row=2, values_only=True):
            rok, dop, kwota = row
            cur.execute("INSERT INTO doplaty_history(username,rok,doplata,kwota) VALUES (?,?,?,?)",
                        (self.username, rok, dop, kwota))
        conn.commit()
        conn.close()
        messagebox.showinfo("Import", "Dane z Excela zaimportowane.")

    # ------------------------------
    # 10. Uprawy (z uprawy.log)
    def build_tab_uprawy(self):
        tb.Label(self.tab_uprawy, text="Uprawy (z pliku uprawy.log):", font="-size 12 -weight bold").pack(pady=10)
        if not os.path.exists(UPRAWY_LOG):
            tb.Label(self.tab_uprawy, text="Brak pliku uprawy.log").pack(pady=10)
            return
        with open(UPRAWY_LOG, "r", encoding="utf-8") as f:
            for line in f:
                tb.Label(self.tab_uprawy, text=line.strip()).pack(anchor="w", padx=20)

    # ------------------------------
    # 11. Analizy – ranking, trendy, prognozy
    def build_tab_analizy(self):
        tb.Label(self.tab_analizy, text="Analizy dopłat:", font="-size 12 -weight bold").pack(pady=10)

        tb.Button(self.tab_analizy, text="Ranking dopłat", bootstyle="info", command=self.show_ranking).pack(pady=5)
        tb.Button(self.tab_analizy, text="Trendy", bootstyle="primary", command=self.show_trendy).pack(pady=5)
        tb.Button(self.tab_analizy, text="Prognoza", bootstyle="success", command=self.show_prognoza).pack(pady=5)

        self.analizy_frame = tb.Frame(self.tab_analizy)
        self.analizy_frame.pack(fill="both", expand=True)

    def show_ranking(self):
        for widget in self.analizy_frame.winfo_children():
            widget.destroy()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT doplata,SUM(kwota) FROM doplaty_history WHERE username=? GROUP BY doplata ORDER BY SUM(kwota) DESC LIMIT 5", (self.username,))
        rows = cur.fetchall()
        conn.close()
        tb.Label(self.analizy_frame, text="Ranking dopłat:", font="-size 12 -weight bold").pack()
        for dop, suma in rows:
            tb.Label(self.analizy_frame, text=f"{dop}: {suma} zł").pack(anchor="w", padx=20)

    def show_trendy(self):
        for widget in self.analizy_frame.winfo_children():
            widget.destroy()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT rok,SUM(kwota) FROM doplaty_history WHERE username=? GROUP BY rok ORDER BY rok", (self.username,))
        rows = cur.fetchall()
        conn.close()
        tb.Label(self.analizy_frame, text="Trendy roczne:", font="-size 12 -weight bold").pack()
        for rok, suma in rows:
            tb.Label(self.analizy_frame, text=f"{rok}: {suma} zł").pack(anchor="w", padx=20)

    def show_prognoza(self):
        for widget in self.analizy_frame.winfo_children():
            widget.destroy()
        conn = sqlite3.connect(DOPLATY_DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT AVG(kwota) FROM doplaty_history WHERE username=?", (self.username,))
        avg = cur.fetchone()[0]
        conn.close()
        tb.Label(self.analizy_frame, text="Prognoza na kolejny rok:", font="-size 12 -weight bold").pack()
        tb.Label(self.analizy_frame, text=f"Średnia kwota dopłat: {avg or 0:.2f} zł").pack(anchor="w", padx=20)

    # ------------------------------
    # 12. Backup i ustawienia
    def build_tab_backup(self):
        tb.Label(self.tab_backup, text="Backup i ustawienia:", font="-size 12 -weight bold").pack(pady=10)
        tb.Button(self.tab_backup, text="Backup bazy", bootstyle="success", command=self.backup_db).pack(pady=5)
        tb.Button(self.tab_backup, text="Przywróć bazę", bootstyle="danger", command=self.restore_db).pack(pady=5)

    def backup_db(self):
        fname = filedialog.asksaveasfilename(defaultextension=".db")
        if not fname: return
        import shutil
        shutil.copy(DOPLATY_DB_FILE, fname)
        messagebox.showinfo("Backup", "Baza została skopiowana.")

    def restore_db(self):
        fname = filedialog.askopenfilename(filetypes=[("Baza SQLite", "*.db")])
        if not fname: return
        import shutil
        shutil.copy(fname, DOPLATY_DB_FILE)
        messagebox.showinfo("Przywracanie", "Baza została przywrócona.")

# =============================
#   MAIN
# =============================
if __name__ == "__main__":
    u = load_session_username()
    if not u:
        messagebox.showerror("Brak sesji", "Zaloguj się przez panel główny.")
        sys.exit(1)

    app = DoplatyApp(u)

    # budowanie wszystkich zakładek
    app.build_tab_lista()
    app.build_tab_stawki()
    app.build_tab_terminy()
    app.build_tab_checklist()
    app.build_tab_calc()
    app.build_tab_historia()
    app.build_tab_alerty()
    app.build_tab_raport()
    app.build_tab_excel()
    app.build_tab_uprawy()
    app.build_tab_analizy()
    app.build_tab_backup()

    app.mainloop()
